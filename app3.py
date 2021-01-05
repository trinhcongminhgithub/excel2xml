import dttype as dt
import xml.etree.ElementTree as ET
import openpyxl as pyxl
import json
import random
codes = {
    "Heart beat":{
        "code":"8867-4",
        "display":"Heart rate"
    },
    "Body temperature":{
        "code":"8310-5",
        "display":"Body temperature"
    },
    "Diabetes":{
        "code":"66152-0",
        "display":"Doctor or health care professional ever told you that you have diabetes"
    },
    "Systolic blood pressure":{
        "code":"8480-6",
        "display":"Systolic blood pressure"   
    },
    "Diastolic blood pressure":{
        "code":"8462-4",
        "display":"Diastolic blood pressure"
    },
    "Respiratory rate":{
        "code":"9279-1",
        "display":"Respiratory rate"
    },
    "Body weight":{
        "code":"29463-7",
        "display":"Body weight"
    }
}
valueQuan = {
    "Heart beat":{
        "code":"{Beats}/min",
        "display":"Beats / minute"
        },
    "Body temperature":{
        "code":"Cel",
        "display":"degree Celcius"
    },
    "Systolic blood pressure":{
        "code":"mm[Hg]",
        "display":"MilliMetersOfMercury"
    },
    "Diastolic blood pressure":{
        "code":"mm[Hg]",
        "display":"MilliMetersOfMercury"
    },
    "Respiratory rate":{
        "code":"{Breaths}/min",
        "display":"Breaths / minute"
    },
    "Body weight":{
        "code":"kg",
        "display":"kilogram"
    }
}

wb = pyxl.load_workbook("DC_TrinhCongMinh_thongke.xlsx")

sheet = wb["Sheet1"]
m_row = sheet.max_row

contents = []
for i in range(1,m_row+1):
    cell = sheet.cell(row=i,column=4)
    if cell.value == None:
        continue
    cell_element = cell.value.split('.')
    if cell_element[0] == "Observation":
        cell_element.insert(1,i)
        print(cell_element)
        contents.append(cell_element[1:])
i = 0
ids = []


for content in contents:
    root = ET.Element('Observation')
    tree = ET.ElementTree(root)
    root.set("xmlns","http://hl7.org/fhir")
    identifier_res = ET.SubElement(root, 'identifier')
    identifier = random.randint(0,10001)
    while(identifier in ids):
        identifier = random.randint(0,10001)
    ids.append(identifier)
    dt.identifier_type(identifier_res, "http://localhost",identifier)

    cell = sheet.cell(row=content[0], column=2)
    new_content = json.loads(content[1])
    for key in new_content.keys():
        if key == "basedOn":
            basedOn_res = ET.SubElement(root, 'basedOn')
        elif key == "partOf":
            partOf_res = ET.SubElement(root, 'partOf')
        elif key == "status":
            status_res = ET.SubElement(root, 'status')
            status_res.set('value',new_content[key])
        elif key == "category":
            category_res = ET.SubElement(root, 'category')
            coding = ET.SubElement(category_res, 'coding')
            system = "http://terminology.hl7.org/CodeSystem/observation-category"
            cat_code = new_content[key]
            dt.coding_type(coding, system, cat_code)
        elif key == "code":
            code_res = ET.SubElement(root, 'code')
            coding = ET.SubElement(code_res, 'coding')
            system = "http://loinc.org"
            code_code = codes[cell.value.split(':')[0]]['code']
            display = codes[cell.value.split(':')[0]]['display']
            dt.coding_type(coding, system, code_code, display)
        elif key == "subject":
            subject_res = ET.SubElement(root, 'subject')
        elif key == "focus":
            focus_res = ET.SubElement(root, 'focus')
        elif key == "encounter":
            encounter_res = ET.SubElement(root, 'encounter')
        # elif key == "effective{}".format(content[2]):
        #     effective_res = ET.SubElement(root, 'effective{}'.format(content[2]))
        elif key == "issued":
            issued_res = ET.SubElement(root, 'issued')
        elif key == "performer":
            performer_res = ET.SubElement(root, 'performer')
        elif key == "valueQuantity":
            value_res = ET.SubElement(root, 'valueQuantity')
            value = cell.value.split(':')[1].split(' ')[0]
            unit = cell.value.split(':')[1].split(' ')[1]
            system = "http://unitsofmeasure.org"
            val_code = valueQuan[cell.value.split(':')[0]]['code']
            dt.quantity_type(value_res, value, unit, system, val_code)
        elif key == "dataAbsentReason":
            dataAbsentReason_res = ET.SubElement(root, 'dataAbsentReason')
            coding = ET.SubElement(dataAbsentReason_res, 'coding')
            system = "http://terminology.hl7.org/CodeSystem/data-absent-reason"
            absent_code = new_content[key]
            dt.coding_type(coding, system, absent_code)
        elif key == "interpretation":
            interpretation_res = ET.SubElement(root, 'interpretation')
        elif key == "note":
            note_res = ET.SubElement(root, 'note')
        elif key == "bodySite":
            bodySite_res = ET.SubElement(root, 'bodySite')
        elif key == "method":
            method_res = ET.SubElement(root, 'method')
        elif key == "specimen":
            specimen_res = ET.SubElement(root, 'specimen')
        elif key == "device":
            device_res = ET.SubElement(root, 'device')
        elif key == "referenceRange":
            referenceRange_res = ET.SubElement(root, 'referenceRange')
        elif key == "hasMember":
            hasMember_res = ET.SubElement(root, 'hasMember')
        elif key == "derivedFrom":
            derivedFrom_res = ET.SubElement(root, 'derivedFrom')
        elif key == "component":
            component_res = ET.SubElement(root, 'component')
    with open("Observation{}.xml".format(i),"wb") as f:
        tree.write(f)
    i += 1