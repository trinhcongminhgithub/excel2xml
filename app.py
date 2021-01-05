import dttype as dt
import xml.etree.ElementTree as ET
import openpyxl as pyxl
import random
import json
import uuid
root = ET.Element('Patient')
tree = ET.ElementTree(root)
wb = pyxl.load_workbook("DC_TrinhCongMinh_thongke.xlsx")
root.set("xmlns","http://hl7.org/fhir")
sheet = wb["Sheet1"]
m_row = sheet.max_row



contents = []
for i in range(1,m_row+1):
    cell = sheet.cell(row=i,column=4)
    if cell.value == None:
        continue
    cell_element = cell.value.split('.')
    if cell_element[0] == "Patient":
        cell_element.insert(1,i)
        contents.append(cell_element[1:])

my_own_order = ['identifier', 'active', 'name', 'telecom', 'gender', 'birthDate', 'deceaseddateTime', 'address', 'maritalStatus', 'multipleBirth', 'photo', 'contact', 'communication', 'generalPractitioner', 'managingOrganization', 'link']
order = {key: i for i, key in enumerate(my_own_order)}
print(order)
print(contents)
new_contents = sorted(contents, key=lambda x: order['{}'.format(x[1])])
print(new_contents)


identifier_res = ET.SubElement(root, 'identifier')
identifier = random.randint(0,10001)
dt.identifier_type(identifier_res, "http://localhost",identifier)


for content in new_contents:
    cell = sheet.cell(row=content[0], column=2)
    if content[1] == "name":
        name = ET.SubElement(root, 'name')
        dt.name_type(name, cell.value)
    elif content[1] == "identifier":
        identifier_res = ET.SubElement(root, 'identifier')
        use = content[2]
        system = "http://localhost"
        if use == "official":
            identifier = cell.value
        dt.identifier_type(identifier_res, system, cell.value, use)
    elif content[1] == "address":
        use = content[2]
        address = ET.SubElement(root, 'address')
        dt.address_type(address, cell.value, use=use)
    elif content[1] == "contact":
        contact = ET.SubElement(root, "contact")
        relationship = ET.SubElement(contact, "relationship")
        system = "http://terminology.hl7.org/CodeSystem/v2-0131"
        code = "C"
        parse = cell.value.split('; ')
        coding = ET.SubElement(relationship, 'coding')
        dt.coding_type(coding, system, code)
        txt = ET.SubElement(relationship, "text value=\"{}\"".format(parse[2]))
        name = ET.SubElement(contact, 'name')
        dt.name_type(name, parse[0])
        print(parse[1])
        address = ET.SubElement(contact, 'address')
        dt.address_type(address, parse[1])
    else:
        ET.SubElement(root, content[1]+ ' value=\"{}\"'.format(cell.value))

with open("Patient.xml","wb") as f:
    tree.write(f)

# def patient(resource):
#     wb = pyxl.load_workbook("DC_TrinhCongMinh_thongke.xlsx")
#     sheet = wb["Sheet1"]
#     m_row = sheet.max_row
#     contents = []
#     for i in range(1,m_row+1):
#        cell = sheet.cell(row=i,column=4)
#        if cell.value == None:
#            continue
#        cell_element = cell.value.split('.')
#        if cell_element[0] == "Patient":
#            cell_element.insert(1,i)
#            contents.append(cell_element[1:])

#     my_own_order = ['identifier', 'active', 'name', 'telecom', 'gender', 'birthDate', 'deceaseddateTime', 'address', 'maritalStatus', 'multipleBirth', 'photo', 'contact', 'communication', 'generalPractitioner', 'managingOrganization', 'link']
#     order = {key: i for i, key in enumerate(my_own_order)}
#     print(order)
#     print(contents)
#     new_contents = sorted(contents, key=lambda x: order['{}'.format(x[1])])
#     print(new_contents)
#     identifier_res = ET.SubElement(resource, 'identifier')
#     identifier = random.randint(0,10001)
#     identifier_type(identifier_res, "http://localhost",identifier)
#     for content in new_contents:
#         cell = sheet.cell(row=content[0], column=2)
#         if content[1] == "name":
#             name = ET.SubElement(resource, 'name')
#             name_type(name, cell.value)
#         # elif content[1] == "identifier":
#         #     identifier_res = ET.SubElement(resource, 'identifier')
#         #     use = content[2]
#         #     system = "http://localhost"
#         #     identifier = random.randint(0,10001)
#             # if use == "official":
#             #     identifier = cell.value
#             #     identifier_type(identifier_res, system, cell.value, use)
#         elif content[1] == "address":
#             use = content[2]
#             address = ET.SubElement(resource, 'address')
#             address_type(address, cell.value, use=use)
#         elif content[1] == "contact":
#             contact = ET.SubElement(resource, "contact")
#             relationship = ET.SubElement(contact, "relationship")
#             system = "http://terminology.hl7.org/CodeSystem/v2-0131"
#             code = "C"
#             parse = cell.value.split('; ')
#             coding = ET.SubElement(relationship, 'coding')
#             coding_type(coding, system, code)
#             ET.SubElement(relationship, "text value=\"{}\"".format(parse[2]))
#             name = ET.SubElement(contact, 'name')
#             name_type(name, parse[0])
#             print(parse[1])
#             address = ET.SubElement(contact, 'address')
#             address_type(address, parse[1])
#         else:
#             ET.SubElement(resource, content[1]+ ' value=\"{}\"'.format(cell.value))
#     return identifier

# def encounter(resource):
#     reasons = {"Motor vehicle traffic accident":"214031005"}
#     codes = {"Khoa điều trị":"IMP"}
#     khoa = {"khoa A":"1", "khoa B": "2"}
#     wb = pyxl.load_workbook("DC_TrinhCongMinh_thongke.xlsx")
#     sheet = wb["Sheet1"]
#     m_row = sheet.max_row

#     contents = []
#     for i in range(1,m_row+1):
#         cell = sheet.cell(row=i,column=4)
#         if cell.value == None:
#             continue
#         cell_element = cell.value.split('.')
#         if cell_element[0] == "Encounter":
#             cell_element.insert(1,i)
#             contents.append(cell_element[1:])

#     identifier_res = ET.SubElement(resource, 'identifier')
#     identifier = random.randint(0,10001)
#     identifier_type(identifier_res, "http://localhost",identifier)

#     for content in contents:
#         cell = sheet.cell(row=content[0], column=2)
#         if content[1] == "class":
#             class_res = ET.SubElement(resource, 'class')
#             system = "http://terminology.hl7.org/CodeSystem/v3-ActCode"
#             code = codes[cell.value]
#             print(code)
#             coding_type(class_res, system, code)
#         elif content[1] == 'statusHistory':
#             stat_his = ET.SubElement(resource, 'statusHistory')
#             # status = ""
#             # period = ""
#             # period_res = ET.SubElement(stat_his,'period')
#             # app.period_type(stat_his, period) 
#         elif content[1] == "classHistory":
#             class_his = ET.SubElement(resource, 'classHistory')
#             # system = "http://terminology.hl7.org/CodeSystem/v3-ActCode"
#             # code = "SS"
#             # period = ""
#             # period_res =ET.SubElement(class_his, 'period')
#             # class_res = ET.SubElement(class_his, 'class')
#             # app.coding_type(class_res, system, code)
#             # app.period_type(stat_his, period)
#         elif content[1] == "type":
#             type_res = ET.SubElement(resource, 'type')
#             # system = ""
#             # code = ""
#             # app.coding_type(type_res, system, code)
#         elif content[1] == "serviceType":
#             serviceType_res = ET.SubElement(resource, 'serviceType')
#         elif content[1] == "priority":
#             priority_res = ET.SubElement(resource, 'priority')
#         elif content[1] == "subject":
#             subject_res = ET.SubElement(resource, 'subject')
#         elif content[1] == "episodeOfCare":
#             episodeOfCare_res = ET.SubElement(resource, 'episodeofCare')    
#         elif content[1] == "basedOn":
#             basedOn_res = ET.SubElement(resource, 'basedOn')
#         elif content[1] == "participant":
#             participant_res = ET.SubElement(resource, 'participant')
#         elif content[1] == "appointment":
#             appointment_res = ET.SubElement(resource, 'appointment')
#         elif content[1] == "period":
#             if resource.find('period') == None:
#                 period_res  = ET.SubElement(resource, 'period') 
#             if content[2] == "start":
#                 start = cell.value
#                 period_type(period_res, start=start)      
#             elif content[2] == "end":
#                 end = cell.value
#                 period_type(period_res,end=end)        
#         elif content[1] == "length":
#             length_res = ET.SubElement(resource, 'length')
#             length = cell.value.split(' ')
#             system = "http://unitsofmeasure.org"
#             duration_type(length_res, length[0], length[1], system, length[1][0]) 
#         elif content[1] == "reasonCode":
#             reasonCode_res = ET.SubElement(resource, 'reasonCode')
#             if cell.value in reasons:
#                 system = "http://snomed.info/sct"
#                 code = reasons[cell.value]
#                 coding_type(reasonCode_res, system, code, cell.value)
#             ET.SubElement(reasonCode_res, 'text value=\"{}\"'.format(cell.value))
#         elif content[1] == "reasonReference":
#             reasonReference_res = ET.SubElement(resource, 'reasonReference')
#         elif content[1] == "diagnosis":
#             diagnosis_res = ET.SubElement(resource, 'diagnosis')
#         elif content[1] == "account":
#             account_res = ET.SubElement(resource, 'account')
#         elif content[1] == "hospitalization":
#             if resource.find('hospitalization') == None:
#                 hospitalization_res = ET.SubElement(resource, 'hospitalization')
#         elif content[1] == "location":
#             location_res = ET.SubElement(resource,'location')
#             location = cell.value
#         elif content[1] == "serviceProvider":
#             serviceProvider_res = ET.SubElement(resource, 'serviceProvider')
#         elif content[1] == "partOf":
#             partOf_res = ET.SubElement(resource, 'partOf')

# def observation(resource):
#     codes = {
#     "Heart beat":{
#         "code":"8867-4",
#         "display":"Heart rate"
#     },
#     "Body temperature":{
#         "code":"8310-5",
#         "display":"Body temperature"
#     },
#     "Diabetes":{
#         "code":"66152-0",
#         "display":"Doctor or health care professional ever told you that you have diabetes"
#     },
#     "Systolic blood pressure":{
#         "code":"8480-6",
#         "display":"Systolic blood pressure"   
#     },
#     "Diastolic blood pressure":{
#         "code":"8462-4",
#         "display":"Diastolic blood pressure"
#     },
#     "Respiratory rate":{
#         "code":"9279-1",
#         "display":"Respiratory rate"
#     },
#     "Body weight":{
#         "code":"29463-7",
#         "display":"Body weight"
#     }
#     }
#     valueQuan = {
#     "Heart beat":{
#         "code":"{Beats}/min",
#         "display":"Beats / minute"
#         },
#     "Body temperature":{
#         "code":"Cel",
#         "display":"degree Celcius"
#     },
#     "Systolic blood pressure":{
#         "code":"mm[Hg]",
#         "display":"MilliMetersOfMercury"
#     },
#     "Diastolic blood pressure":{
#         "code":"mm[Hg]",
#         "display":"MilliMetersOfMercury"
#     },
#     "Respiratory rate":{
#         "code":"{Breaths}/min",
#         "display":"Breaths / minute"
#     },
#     "Body weight":{
#         "code":"kg",
#         "display":"kilogram"
#     }
#     }

#     wb = pyxl.load_workbook("DC_TrinhCongMinh_thongke.xlsx")

#     sheet = wb["Sheet1"]
#     m_row = sheet.max_row

#     contents = []
#     for i in range(1,m_row+1):
#         cell = sheet.cell(row=i,column=4)
#         if cell.value == None:
#             continue
#         cell_element = cell.value.split('.')
#         if cell_element[0] == "Observation":
#             cell_element.insert(1,i)
#             print(cell_element)
#             contents.append(cell_element[1:])
#     i = 0
#     ids = []


#     for content in contents:
#         observation_res = ET.SubElement(resource, 'Observation')
#         identifier_res = ET.SubElement(observation_res, 'identifier')
#         identifier = random.randint(0,10001)
#         while(identifier in ids):
#             identifier = random.randint(0,10001)
#         ids.append(identifier)
#         identifier_type(identifier_res, "http://localhost",identifier)
#         cell = sheet.cell(row=content[0], column=2)
#         new_content = json.loads(content[1])
#         for key in new_content.keys():
#             if key == "basedOn":
#                 basedOn_res = ET.SubElement(observation_res, 'basedOn')
#             elif key == "partOf":
#                 partOf_res = ET.SubElement(observation_res, 'partOf')
#             elif key == "status":
#                 status_res = ET.SubElement(observation_res, 'status')
#                 status_res.set('value',new_content[key])
#             elif key == "category":
#                 category_res = ET.SubElement(observation_res, 'category')
#                 coding = ET.SubElement(category_res, 'coding')
#                 system = "http://terminology.hl7.org/CodeSystem/observation-category"
#                 cat_code = new_content[key]
#                 coding_type(coding, system, cat_code)
#             elif key == "code":
#                 code_res = ET.SubElement(observation_res, 'code')
#                 coding = ET.SubElement(code_res, 'coding')
#                 system = "http://loinc.org"
#                 code_code = codes[cell.value.split(':')[0]]['code']
#                 display = codes[cell.value.split(':')[0]]['display']
#                 coding_type(coding, system, code_code, display)
#             elif key == "subject":
#                 subject_res = ET.SubElement(observation_res, 'subject')
#             elif key == "focus":
#                 focus_res = ET.SubElement(observation_res, 'focus')
#             elif key == "encounter":
#                 encounter_res = ET.SubElement(observation_res, 'encounter')
#             # elif key == "effective{}".format(content[2]):
#             #     effective_res = ET.SubElement(root, 'effective{}'.format(content[2]))
#             elif key == "issued":
#                 issued_res = ET.SubElement(observation_res, 'issued')
#             elif key == "performer":
#                 performer_res = ET.SubElement(observation_res, 'performer')
#             elif key == "valueQuantity":
#                 value_res = ET.SubElement(observation_res, 'valueQuantity')
#                 value = cell.value.split(':')[1].split(' ')[0]
#                 unit = cell.value.split(':')[1].split(' ')[1]
#                 system = "http://unitsofmeasure.org"
#                 val_code = valueQuan[cell.value.split(':')[0]]['code']
#                 quantity_type(value_res, value, unit, system, val_code)
#             elif key == "dataAbsentReason":
#                 dataAbsentReason_res = ET.SubElement(observation_res, 'dataAbsentReason')
#                 coding = ET.SubElement(dataAbsentReason_res, 'coding')
#                 system = "http://terminology.hl7.org/CodeSystem/data-absent-reason"
#                 absent_code = new_content[key]
#                 coding_type(coding, system, absent_code)
#             elif key == "interpretation":
#                 interpretation_res = ET.SubElement(observation_res, 'interpretation')
#             elif key == "note":
#                 note_res = ET.SubElement(observation_res, 'note')
#             elif key == "bodySite":
#                 bodySite_res = ET.SubElement(observation_res, 'bodySite')
#             elif key == "method":
#                 method_res = ET.SubElement(observation_res, 'method')
#             elif key == "specimen":
#                 specimen_res = ET.SubElement(observation_res, 'specimen')
#             elif key == "device":
#                 device_res = ET.SubElement(observation_res, 'device')
#             elif key == "referenceRange":
#                 referenceRange_res = ET.SubElement(observation_res, 'referenceRange')
#             elif key == "hasMember":
#                 hasMember_res = ET.SubElement(observation_res, 'hasMember')
#             elif key == "derivedFrom":
#                 derivedFrom_res = ET.SubElement(observation_res, 'derivedFrom')
#             elif key == "component":
#                 component_res = ET.SubElement(observation_res, 'component')
#         i += 1
#     return identifier