import dttype as dt
import xml.etree.ElementTree as ET
import openpyxl as pyxl
import random

reasons = {"Motor vehicle traffic accident":"214031005"}
codes = {"Khoa điều trị":"IMP"}
khoa = {"khoa A":"1", "khoa B": "2"}
root = ET.Element('Encounter')
tree = ET.ElementTree(root)
wb = pyxl.load_workbook("DC_TrinhCongMinh_thongke.xlsx")
root.set("xmlns","http://hl7.org/fhir")
sheet = wb["Sheet1"]
m_row = sheet.max_row

contents = []
for i in range(1,m_row+1):
    cell = sheet.cell(row=i,column=4)
    if cell.value == None:
        continue
    cell_element = cell.value.split('.')
    if cell_element[0] == "Encounter":
        cell_element.insert(1,i)
        contents.append(cell_element[1:])

identifier_res = ET.SubElement(root, 'identifier')
identifier = random.randint(0,10001)
dt.identifier_type(identifier_res, "http://localhost",identifier)

for content in contents:
    cell = sheet.cell(row=content[0], column=2)
    if content[1] == "class":
        class_res = ET.SubElement(root, 'class')
        system = "http://terminology.hl7.org/CodeSystem/v3-ActCode"
        code = codes[cell.value]
        print(code)
        dt.coding_type(class_res, system, code)
    elif content[1] == 'statusHistory':
        stat_his = ET.SubElement(root, 'statusHistory')
        # status = ""
        # period = ""
        # period_res = ET.SubElement(stat_his,'period')
        # app.period_type(stat_his, period)
    elif content[1] == "classHistory":
        class_his = ET.SubElement(root, 'classHistory')
        # system = "http://terminology.hl7.org/CodeSystem/v3-ActCode"
        # code = "SS"
        # period = ""
        # period_res =ET.SubElement(class_his, 'period')
        # class_res = ET.SubElement(class_his, 'class')
        # app.coding_type(class_res, system, code)
        # app.period_type(stat_his, period)
    elif content[1] == "type":
        type_res = ET.SubElement(root, 'type')
        # system = ""
        # code = ""
        # app.coding_type(type_res, system, code)
    elif content[1] == "serviceType":
        serviceType_res = ET.SubElement(root, 'serviceType')
    elif content[1] == "priority":
        priority_res = ET.SubElement(root, 'priority')
    elif content[1] == "subject":
        subject_res = ET.SubElement(root, 'subject')
    elif content[1] == "episodeOfCare":
        episodeOfCare_res = ET.SubElement(root, 'episodeofCare')    
    elif content[1] == "basedOn":
        basedOn_res = ET.SubElement(root, 'basedOn')
    elif content[1] == "participant":
        participant_res = ET.SubElement(root, 'participant')
    elif content[1] == "appointment":
        appointment_res = ET.SubElement(root, 'appointment')
    elif content[1] == "period":
        if root.find('period') == None:
            period_res  = ET.SubElement(root, 'period') 
        if content[2] == "start":
            start = cell.value
            dt.period_type(period_res, start=start)      
        elif content[2] == "end":
            end = cell.value
            dt.period_type(period_res,end=end)        
    elif content[1] == "length":
        length_res = ET.SubElement(root, 'length')
        length = cell.value.split(' ')
        system = "http://unitsofmeasure.org"
        dt.duration_type(length_res, length[0], length[1], system, length[1][0]) 
    elif content[1] == "reasonCode":
        reasonCode_res = ET.SubElement(root, 'reasonCode')
        if cell.value in reasons:
            system = "http://snomed.info/sct"
            code = reasons[cell.value]
            dt.coding_type(reasonCode_res, system, code, cell.value)
        ET.SubElement(reasonCode_res, 'text value=\"{}\"'.format(cell.value))


    elif content[1] == "reasonReference":
        reasonReference_res = ET.SubElement(root, 'reasonReference')
    elif content[1] == "diagnosis":
        diagnosis_res = ET.SubElement(root, 'diagnosis')
    elif content[1] == "account":
        account_res = ET.SubElement(root, 'account')
    elif content[1] == "hospitalization":
        if root.find('hospitalization') == None:
            hospitalization_res = ET.SubElement(root, 'hospitalization')
    elif content[1] == "location":
        location_res = ET.SubElement(root,'location')
        location = cell.value
    elif content[1] == "serviceProvider":
        serviceProvider_res = ET.SubElement(root, 'serviceProvider')
    elif content[1] == "partOf":
        partOf_res = ET.SubElement(root, 'partOf')


